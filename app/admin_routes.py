from flask import (
    Blueprint,
    render_template,
    redirect,
    url_for,
    flash,
    session,
    current_app,
    abort,
)
import flask
from functools import wraps
from datetime import datetime, timezone
from pathlib import Path
from uuid import uuid4
import mimetypes

from werkzeug.utils import secure_filename

from . import db, csrf
from . import email_utils

# Alias retained for compatibility with tests that monkeypatch the function.
send_email = email_utils.send_email
from .template_utils import render_template_string
from .forms import (
    CoachForm,
    TrainingForm,
    LoginForm,
    ImportTrainingsForm,
    LocationForm,
    SettingsForm,
)
from .models import Coach, Training, Location, EmailSettings, StoredFile

admin_bp = Blueprint("admin", __name__)


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("admin_logged_in"):
            return redirect(url_for("admin.login"))
        return view(*args, **kwargs)

    return wrapped


@admin_bp.route("/")
def admin_root():
    """Redirect bare /admin to the trainings view."""
    return redirect(url_for("admin.manage_trainings"))


@admin_bp.route("/login", methods=["GET", "POST"])
def login():
    form = LoginForm()

    if form.validate_on_submit():
        stored_password = current_app.config.get("ADMIN_PASSWORD")
        if not stored_password:
            flash("Hasło administratora nie jest skonfigurowane.", "danger")
        elif form.password.data == stored_password:
            session["admin_logged_in"] = True
            flash("Zalogowano jako administrator.", "success")
            return redirect(url_for("admin.manage_trainings"))
        else:
            flash("Nieprawidłowe hasło.", "danger")

    return render_template("admin/login.html", form=form)


@admin_bp.route("/logout")
def logout():
    session.pop("admin_logged_in", None)
    flash("Wylogowano.", "info")
    return redirect(url_for("admin.login"))


@admin_bp.route("/trainers", methods=["GET", "POST"])
@login_required
def manage_trainers():
    form = CoachForm()
    coaches = Coach.query.order_by(Coach.last_name).all()

    if form.validate_on_submit():
        new_coach = Coach(
            first_name=form.first_name.data.strip(),
            last_name=form.last_name.data.strip(),
            phone_number=form.phone_number.data.strip(),
        )
        db.session.add(new_coach)
        db.session.commit()
        flash("Dodano nowego trenera.", "success")
        return redirect(url_for("admin.manage_trainers"))

    return render_template("admin/trainers.html", form=form, coaches=coaches)


@admin_bp.route("/trainers/edit/<int:coach_id>", methods=["GET", "POST"])
@login_required
def edit_trainer(coach_id):
    coach = db.session.get(Coach, coach_id)
    if coach is None:
        abort(404)
    form = CoachForm(obj=coach)

    if form.validate_on_submit():
        coach.first_name = form.first_name.data.strip()
        coach.last_name = form.last_name.data.strip()
        coach.phone_number = form.phone_number.data.strip()
        db.session.commit()
        flash("Zaktualizowano dane trenera.", "success")
        return redirect(url_for("admin.manage_trainers"))

    return render_template("admin/edit_trainer.html", form=form, coach=coach)


@admin_bp.route("/trainers/<int:coach_id>/delete", methods=["POST"])
@login_required
def delete_trainer(coach_id):
    coach = db.session.get(Coach, coach_id)
    if coach is None:
        abort(404)
    if coach.trainings:
        flash(
            "Nie można usunąć trenera powiązanego z treningami.",
            "warning",
        )
        return redirect(url_for("admin.manage_trainers"))

    db.session.delete(coach)
    db.session.commit()
    flash("Trener został usunięty.", "info")
    return redirect(url_for("admin.manage_trainers"))


@admin_bp.route("/locations", methods=["GET", "POST"])
@login_required
def manage_locations():
    form = LocationForm()
    locations = Location.query.order_by(Location.name).all()

    if form.validate_on_submit():
        new_location = Location(name=form.name.data.strip())
        db.session.add(new_location)
        db.session.commit()
        flash("Dodano nowe miejsce.", "success")
        return redirect(url_for("admin.manage_locations"))

    return render_template(
        "admin/locations.html",
        form=form,
        locations=locations,
    )


@admin_bp.route("/locations/edit/<int:location_id>", methods=["GET", "POST"])
@login_required
def edit_location(location_id):
    location = db.session.get(Location, location_id)
    if location is None:
        abort(404)
    form = LocationForm(obj=location)

    if form.validate_on_submit():
        location.name = form.name.data.strip()
        db.session.commit()
        flash("Zaktualizowano miejsce.", "success")
        return redirect(url_for("admin.manage_locations"))

    return render_template(
        "admin/edit_location.html",
        form=form,
        location=location,
    )


@admin_bp.route("/locations/<int:location_id>/delete", methods=["POST"])
@login_required
def delete_location(location_id):
    location = db.session.get(Location, location_id)
    if location is None:
        abort(404)
    if location.trainings:
        flash("Nie można usunąć miejsca, ponieważ jest używane.", "warning")
        return redirect(url_for("admin.manage_locations"))
    db.session.delete(location)
    db.session.commit()
    flash("Miejsce zostało usunięte.", "info")
    return redirect(url_for("admin.manage_locations"))


@admin_bp.route("/trainings", methods=["GET", "POST"])
@login_required
def manage_trainings():
    form = TrainingForm()
    form.coach_id.choices = [
        (
            c.id,
            f"{c.first_name} {c.last_name}",
        )
        for c in Coach.query.order_by(Coach.last_name).all()
    ]
    form.location_id.choices = [
        (loc.id, loc.name) for loc in Location.query.order_by(Location.name).all()
    ]
    repeat_feedback = session.pop("repeat_feedback", None)
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    trainings_q = Training.query.filter(
        Training.date >= today,
        Training.is_deleted.is_(False),
    ).order_by(Training.date)
    trainings = trainings_q.all()

    # Aggregate recurring trainings into series grouped by weekday, time, coach and location
    day_names = [
        "Poniedziałek",
        "Wtorek",
        "Środa",
        "Czwartek",
        "Piątek",
        "Sobota",
        "Niedziela",
    ]
    series_map = {}
    for training in trainings:
        date = training.date
        weekday = date.weekday()
        time_label = date.strftime("%H:%M")
        key = (
            weekday,
            time_label,
            training.coach_id,
            training.location_id,
        )
        series = series_map.get(key)
        if series is None:
            series_key = (
                f"{weekday}-{date.strftime('%H%M')}-c{training.coach_id}-"
                f"l{training.location_id}"
            )
            series = {
                "series_key": series_key,
                "weekday": weekday,
                "weekday_label": day_names[weekday],
                "time_label": time_label,
                "coach_name": f"{training.coach.first_name} {training.coach.last_name}",
                "location_name": training.location.name,
                "count": 0,
                "first_date": date,
                "last_date": date,
            }
            series_map[key] = series

        series["count"] += 1
        if date < series["first_date"]:
            series["first_date"] = date
        if date > series["last_date"]:
            series["last_date"] = date

    series_summary = sorted(
        series_map.values(),
        key=lambda data: (
            data["weekday"],
            data["time_label"],
            data["coach_name"],
            data["location_name"],
        ),
    )

    trainings_by_month = {}
    for t in trainings:
        month_key = t.date.strftime("%Y-%m")
        trainings_by_month.setdefault(month_key, []).append(t)

    if form.validate_on_submit():
        planned_dates = form.iter_occurrences()
        planned_count = len(planned_dates)
        created = 0
        conflicts = []

        for candidate in planned_dates:
            conflict = (
                Training.query.filter(
                    Training.date == candidate,
                    Training.location_id == form.location_id.data,
                    Training.is_deleted.is_(False),
                )
                .order_by(Training.id)
                .first()
            )

            if conflict:
                conflicts.append(candidate)
                continue

            new_training = Training(
                date=candidate,
                location_id=form.location_id.data,
                coach_id=form.coach_id.data,
                max_volunteers=form.max_volunteers.data,
            )
            db.session.add(new_training)
            created += 1

        if created:
            db.session.commit()
        else:
            db.session.rollback()

        conflict_strings = [dt.strftime("%Y-%m-%d %H:%M") for dt in conflicts]

        if conflicts and created:
            summary_category = "warning"
        elif conflicts and not created:
            summary_category = "danger"
        else:
            summary_category = "success"

        if planned_count:
            summary_message = (
                f"Dodano {created} z {planned_count} zaplanowanych treningów."
            )
        else:
            summary_message = "Nie udało się zaplanować treningów."

        if conflicts:
            summary_message += (
                f" Pominięto {len(conflicts)} terminów z konfliktem."
            )

        if (
            planned_count == 1
            and created == 1
            and not conflicts
        ):
            flash_message = "Dodano nowy trening."
        else:
            flash_message = summary_message

        flash(flash_message, summary_category)

        session["repeat_feedback"] = {
            "category": summary_category,
            "created": created,
            "planned": planned_count,
            "skipped": conflict_strings,
            "message": summary_message,
        }

        return redirect(url_for("admin.manage_trainings"))

    return render_template(
        "admin/trainings.html",
        form=form,
        trainings_by_month=trainings_by_month,
        repeat_feedback=repeat_feedback,
        series_summary=series_summary,
    )


@admin_bp.route("/trainings/edit/<int:training_id>", methods=["GET", "POST"])
@login_required
def edit_training(training_id):
    training = db.session.get(Training, training_id)
    if training is None:
        abort(404)
    if training.is_deleted:
        abort(404)
    form = TrainingForm(obj=training)
    form.coach_id.choices = [
        (c.id, f"{c.first_name} {c.last_name}")
        for c in Coach.query.order_by(Coach.last_name).all()
    ]
    form.location_id.choices = [
        (loc.id, loc.name) for loc in Location.query.order_by(Location.name).all()
    ]
    if form.validate_on_submit():
        training.date = form.date.data
        training.location_id = form.location_id.data
        training.coach_id = form.coach_id.data
        training.max_volunteers = form.max_volunteers.data
        db.session.commit()
        flash("Zaktualizowano trening.", "success")
        return redirect(url_for("admin.manage_trainings"))
    return render_template(
        "admin/edit_training.html",
        form=form,
        training=training,
    )


@admin_bp.route("/trainings/<int:training_id>/cancel", methods=["POST"])
@login_required
def cancel_training(training_id):
    training = db.session.get(Training, training_id)
    if training is None:
        abort(404)
    training.is_canceled = True
    db.session.commit()

    subject = "Trening odwołany"
    settings = db.session.get(EmailSettings, 1)
    body_template = (
        settings.cancellation_template
        if settings and settings.cancellation_template
        else None
    )
    data = {
        "date": training.date.strftime("%Y-%m-%d %H:%M"),
        "location": training.location.name,
        "logo": url_for("static", filename="logo.png", _external=True),
    }
    if body_template:
        html_body = render_template_string(body_template, data)
    else:
        html_body = f"Trening {data['date']} w {data['location']} został odwołany."
    recipients = [b.volunteer.email for b in training.bookings]
    if recipients:
        success, error = email_utils.send_email(
            subject, None, recipients, html_body=html_body
        )
        if not success:
            msg = "Nie udało się wysłać e-maila"
            if error:
                msg += f": {error}"
            flash(msg, "danger")

    flash("Trening został oznaczony jako odwołany.", "warning")
    return redirect(url_for("admin.manage_trainings"))


@admin_bp.route("/trainings/<int:training_id>/delete", methods=["POST"])
@login_required
def delete_training(training_id):
    training = db.session.get(Training, training_id)
    if training is None:
        abort(404)
    training.is_deleted = True
    db.session.commit()
    flash("Trening został usunięty.", "info")
    return redirect(url_for("admin.manage_trainings"))


@admin_bp.route("/trainings/series/<series_key>/edit")
@login_required
def edit_series(series_key):
    flash("Edycja serii treningów będzie wkrótce dostępna.", "info")
    return redirect(url_for("admin.manage_trainings"))


@admin_bp.route("/trainings/series/<series_key>/delete", methods=["POST"])
@login_required
def delete_series(series_key):
    flash("Usuwanie serii treningów będzie wkrótce dostępne.", "warning")
    return redirect(url_for("admin.manage_trainings"))


@admin_bp.route("/export")
@login_required
def export_excel():
    from io import BytesIO
    from openpyxl import Workbook
    from flask import send_file

    wb = Workbook()
    ws = wb.active
    ws.title = "Treningi"

    header = [
        "Data",
        "Godzina",
        "Miejsce",
        "Trener",
        "Telefon trenera",
        "Wolontariusz 1",
        "Email 1",
        "Wolontariusz 2",
        "Email 2",
    ]
    ws.append(header)

    trainings = Training.query.order_by(Training.date).all()

    for t in trainings:
        bookings = t.bookings[:2]
        v1 = bookings[0].volunteer if len(bookings) > 0 else None
        v2 = bookings[1].volunteer if len(bookings) > 1 else None

        email1 = v1.email if v1 else ""
        email2 = v2.email if v2 else ""

        ws.append(
            [
                t.date.strftime("%Y-%m-%d"),
                t.date.strftime("%H:%M"),
                t.location.name,
                f"{t.coach.first_name} {t.coach.last_name}",
                t.coach.phone_number,
                f"{v1.first_name} {v1.last_name}" if v1 else "",
                email1,
                f"{v2.first_name} {v2.last_name}" if v2 else "",
                email2,
            ]
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"treningi_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    )


@admin_bp.route("/import", methods=["GET", "POST"])
@login_required
def import_excel():
    form = ImportTrainingsForm()

    if form.validate_on_submit():
        from openpyxl import load_workbook

        file = form.file.data
        wb = load_workbook(file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            date_val, time_val, trainer_name, phone, place = row[:5]
            if not (date_val and time_val and phone and place):
                continue

            if isinstance(date_val, datetime):
                date_part = date_val.date()
            else:
                date_part = datetime.strptime(str(date_val), "%Y-%m-%d").date()

            if isinstance(time_val, datetime):
                time_part = time_val.time()
            else:
                time_part = datetime.strptime(str(time_val), "%H:%M").time()

            dt = datetime.combine(date_part, time_part)

            coach = Coach.query.filter_by(phone_number=str(phone).strip()).first()
            if not coach:
                first, *rest = str(trainer_name).strip().split(" ", 1)
                last = rest[0] if rest else ""
                coach = Coach(
                    first_name=first,
                    last_name=last,
                    phone_number=str(phone).strip(),
                )
                db.session.add(coach)
                db.session.commit()

            location = Location.query.filter_by(name=str(place).strip()).first()
            if not location:
                location = Location(name=str(place).strip())
                db.session.add(location)
                db.session.commit()

            training = Training(
                date=dt,
                coach_id=coach.id,
                location_id=location.id,
            )
            db.session.add(training)
        db.session.commit()
        flash("Zaimportowano treningi.", "success")
        return redirect(url_for("admin.manage_trainings"))

    return render_template("admin/import.html", form=form)


@admin_bp.route("/history")
@login_required
def history():
    """List past trainings with volunteer sign-ups."""
    page = flask.request.args.get("page", 1, type=int)
    stmt = (
        db.select(Training)
        .where(
            db.or_(
                Training.date < datetime.now(timezone.utc),
                Training.is_canceled.is_(True),
                Training.is_deleted.is_(True),
            )
        )
        .order_by(Training.date.desc())
    )
    pagination = db.paginate(stmt, page=page, per_page=10)
    return render_template(
        "admin/history.html",
        trainings=pagination.items,
        pagination=pagination,
    )


@admin_bp.route("/history/<int:training_id>/remove", methods=["POST"])
@login_required
def remove_training(training_id):
    """Permanently delete a training and its bookings."""
    training = db.session.get(Training, training_id)
    if training is None:
        abort(404)
    db.session.delete(training)
    db.session.commit()
    flash("Trening został trwale usunięty.", "info")
    return redirect(url_for("admin.history"))


@admin_bp.route("/settings", methods=["GET", "POST"])
@login_required
def settings():
    """Edit email configuration."""
    settings = db.session.get(EmailSettings, 1)
    if not settings:
        settings = EmailSettings(id=1, port=587, encryption="tls")
        db.session.add(settings)
        db.session.commit()

    form = SettingsForm(obj=settings)
    def _has_uploaded_files(field_name: str) -> bool:
        """Return ``True`` when the request contains real file uploads."""

        files = flask.request.files.getlist(field_name)
        for storage in files:
            if storage and getattr(storage, "filename", ""):
                return True
        return False

    if not _has_uploaded_files("registration_files_adult"):
        form.registration_files_adult.data = []
    if not _has_uploaded_files("registration_files_minor"):
        form.registration_files_minor.data = []

    attachments_dir = Path(current_app.instance_path) / "attachments"

    attachment_error_message = (
        "Nie udało się zmodyfikować załączników. Zmiany nie zostały zapisane."
    )

    error_occurred = False

    missing_files_cleaned = False

    def _normalize_file_entries(entries):
        nonlocal error_occurred, missing_files_cleaned
        normalized = []
        for entry in entries:
            if isinstance(entry, dict):
                stored_name = entry.get("stored_name")
                if stored_name:
                    file_path = attachments_dir / stored_name
                    if not file_path.exists():
                        current_app.logger.warning(
                            "Attachment %s referenced in settings but missing on disk",
                            file_path,
                        )
                        missing_files_cleaned = True
                        continue
                normalized.append(entry)
                continue
            if isinstance(entry, int):
                stored_file = db.session.get(StoredFile, entry)
                if not stored_file:
                    current_app.logger.warning(
                        "Stored file with id %s referenced in settings but missing",
                        entry,
                    )
                    continue
                safe_name = secure_filename(stored_file.filename or "")
                if not safe_name:
                    safe_name = f"file_{entry}"
                stored_name = f"legacy_{entry}_{safe_name}"
                try:
                    attachments_dir.mkdir(parents=True, exist_ok=True)
                except OSError:
                    current_app.logger.exception(
                        "Failed to create attachments directory %s", attachments_dir
                    )
                    flash(attachment_error_message, "danger")
                    error_occurred = True
                    normalized.append(
                        {
                            "stored_name": stored_name,
                            "original_name": stored_file.filename,
                            "content_type": stored_file.content_type,
                        }
                    )
                    continue
                target_path = attachments_dir / stored_name
                if not target_path.exists():
                    try:
                        target_path.write_bytes(stored_file.data)
                    except OSError:
                        current_app.logger.warning(
                            "Failed to materialize legacy stored file %s", target_path
                        )
                        continue
                normalized.append(
                    {
                        "stored_name": stored_name,
                        "original_name": stored_file.filename,
                        "content_type": stored_file.content_type,
                    }
                )
                continue
            current_app.logger.warning(
                "Unexpected attachment metadata entry in settings: %r", entry
            )
        return normalized

    existing_adult = _normalize_file_entries(list(settings.registration_files_adult or []))
    existing_minor = _normalize_file_entries(list(settings.registration_files_minor or []))

    if missing_files_cleaned:
        settings.registration_files_adult = existing_adult
        settings.registration_files_minor = existing_minor
        db.session.commit()

    def _choice_label(meta):
        return meta.get("original_name") or meta.get("filename") or meta.get("stored_name")

    form.remove_adult_files.choices = [
        (meta.get("stored_name"), _choice_label(meta))
        for meta in existing_adult
        if meta.get("stored_name")
    ]
    form.remove_minor_files.choices = [
        (meta.get("stored_name"), _choice_label(meta))
        for meta in existing_minor
        if meta.get("stored_name")
    ]

    if form.validate_on_submit():
        settings.server = form.server.data.strip() if form.server.data else None
        settings.port = form.port.data
        settings.encryption = form.encryption.data
        settings.login = form.login.data.strip() if form.login.data else None
        settings.password = form.password.data if form.password.data else None
        settings.sender = form.sender.data.strip()
        settings.registration_template = form.registration_template.data
        settings.cancellation_template = form.cancellation_template.data

        try:
            attachments_dir.mkdir(parents=True, exist_ok=True)
        except OSError:
            current_app.logger.exception(
                "Failed to create attachments directory %s", attachments_dir
            )
            flash(attachment_error_message, "danger")
            error_occurred = True

        def _process_files(uploaded, existing, removals):
            nonlocal error_occurred
            if error_occurred:
                return existing

            updated = list(existing)

            for stored_name in removals:
                if not stored_name:
                    continue
                file_path = attachments_dir / stored_name
                if file_path.exists():
                    try:
                        file_path.unlink()
                    except OSError:
                        current_app.logger.exception(
                            "Failed to delete attachment %s", file_path
                        )
                        flash(attachment_error_message, "danger")
                        error_occurred = True
                        return existing
                updated = [
                    meta
                    for meta in updated
                    if meta.get("stored_name") != stored_name
                ]

            created_files = []

            for storage in uploaded:
                if not storage or not storage.filename:
                    continue
                original_name = storage.filename
                secure_name = secure_filename(original_name)
                if not secure_name:
                    continue
                stored_name = f"{uuid4().hex}_{secure_name}"
                target_path = attachments_dir / stored_name
                try:
                    storage.save(target_path)
                except OSError:
                    current_app.logger.exception(
                        "Failed to save attachment %s", target_path
                    )
                    flash(attachment_error_message, "danger")
                    error_occurred = True
                    for created in created_files:
                        cleanup_path = attachments_dir / created
                        try:
                            if cleanup_path.exists():
                                cleanup_path.unlink()
                        except OSError:
                            current_app.logger.warning(
                                "Failed to clean up attachment %s after error",
                                cleanup_path,
                            )
                    return existing
                created_files.append(stored_name)
                content_type = (
                    storage.mimetype
                    or mimetypes.guess_type(original_name)[0]
                    or "application/octet-stream"
                )
                updated.append(
                    {
                        "stored_name": stored_name,
                        "original_name": original_name,
                        "content_type": content_type,
                    }
                )
            return updated

        if not error_occurred:
            adult_removals = set(form.remove_adult_files.data or [])
            minor_removals = set(form.remove_minor_files.data or [])

            updated_adult = _process_files(
                form.registration_files_adult.data or [],
                existing_adult,
                adult_removals,
            )
            updated_minor = _process_files(
                form.registration_files_minor.data or [],
                existing_minor,
                minor_removals,
            )

        if error_occurred:
            db.session.rollback()
            return render_template("admin/settings.html", form=form)

        settings.registration_files_adult = updated_adult
        settings.registration_files_minor = updated_minor
        db.session.commit()
        flash("Zapisano ustawienia.", "success")
        return redirect(url_for("admin.settings"))

    return render_template("admin/settings.html", form=form)


@admin_bp.route("/settings/test-email", methods=["POST"])
@login_required
def test_email():
    """Send a test email using provided settings without saving."""
    form = SettingsForm()
    form.remove_adult_files.choices = []
    form.remove_minor_files.choices = []
    if form.validate_on_submit():
        recipient_data = form.test_recipient.data
        if not recipient_data:
            flash(
                "Podaj adres e-mail odbiorcy wiadomości testowej.",
                "warning",
            )
        else:
            recipient = recipient_data.strip()
            if not recipient:
                flash(
                    "Podaj adres e-mail odbiorcy wiadomości testowej.",
                    "warning",
                )
            else:
                try:
                    success, error = email_utils.send_email(
                        "Test konfiguracji",
                        "To jest testowa wiadomość.",
                        [recipient],
                        host=form.server.data or None,
                        port=form.port.data,
                        username=form.login.data or None,
                        password=form.password.data or None,
                        encryption=form.encryption.data,
                        sender=form.sender.data or None,
                    )
                    if success:
                        flash("Wysłano wiadomość testową.", "success")
                    else:
                        msg = "Nie udało się wysłać wiadomości testowej"
                        if error:
                            msg += f": {error}"
                        flash(msg, "danger")
                except Exception:  # pragma: no cover - safety net
                    current_app.logger.exception("Failed to send test email")
                    flash("Nie udało się wysłać wiadomości testowej.", "danger")
    else:
        flash("Nie udało się wysłać wiadomości testowej.", "danger")
    return render_template("admin/settings.html", form=form)


@admin_bp.route("/settings/preview/<template>", methods=["GET", "POST"])
@login_required
@csrf.exempt
def preview_template(template):
    settings = db.session.get(EmailSettings, 1)

    if flask.request.method == "POST" and "content" in flask.request.form:
        tpl = flask.request.form.get("content", "")
    else:
        if template not in ("registration", "cancellation"):
            abort(404)
        if settings:
            if template == "registration":
                tpl = settings.registration_template or ""
            else:  # template == "cancellation"
                tpl = settings.cancellation_template or ""
        else:
            tpl = ""

    data = {
        "first_name": "Jan",
        "last_name": "Kowalski",
        "training": "2024-01-01 10:00 w Warszawie",
        "cancel_link": "https://example.com/cancel",
        "date": "2024-01-01 10:00",
        "location": "Warszawa",
        "logo": url_for("static", filename="logo.png", _external=True),
    }
    html = render_template_string(tpl, data)
    return render_template("admin/preview_email.html", html=html)
