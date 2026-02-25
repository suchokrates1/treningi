"""CLI commands for scheduled tasks."""

import click
import secrets
from flask import current_app
from flask.cli import with_appcontext
from datetime import datetime, timezone, timedelta
from zoneinfo import ZoneInfo

from . import db
from .models import Training, Booking, Volunteer, EmailSettings
from .whatsapp_utils import notify_volunteer_reminder, notify_volunteer_reminder_multi, format_phone_display
from .email_utils import send_email
from .template_utils import render_template_string


# Minimum hours since signup before sending reminder
# (to avoid sending reminder right after signup confirmation)
MIN_HOURS_SINCE_SIGNUP = 4


@click.command('send-reminders')
@with_appcontext
def send_reminders_command():
    """Send WhatsApp reminders for tomorrow's trainings (run daily in evening)."""
    tomorrow = datetime.now(timezone.utc).date() + timedelta(days=1)
    tomorrow_start = datetime.combine(tomorrow, datetime.min.time()).replace(tzinfo=timezone.utc)
    tomorrow_end = datetime.combine(tomorrow, datetime.max.time()).replace(tzinfo=timezone.utc)
    
    # Cutoff time: don't send reminders to people who signed up less than MIN_HOURS ago
    signup_cutoff = datetime.now(timezone.utc) - timedelta(hours=MIN_HOURS_SINCE_SIGNUP)

    trainings = Training.query.filter(
        Training.date >= tomorrow_start,
        Training.date <= tomorrow_end,
        Training.is_canceled.is_(False),
        Training.is_deleted.is_(False),
    ).all()

    if not trainings:
        click.echo("No trainings scheduled for tomorrow.")
        return

    sent_count = 0
    failed_count = 0
    skipped_count = 0

    # Group bookings by volunteer so we can send one combined message
    from collections import defaultdict
    volunteer_bookings: dict[int, list] = defaultdict(list)  # vol_id -> [(training, booking)]
    
    for training in trainings:
        for booking in training.bookings:
            volunteer = booking.volunteer
            if not volunteer.phone_number:
                current_app.logger.info(
                    "Volunteer %s %s has no phone number, skipping WhatsApp reminder",
                    volunteer.first_name,
                    volunteer.last_name,
                )
                skipped_count += 1
                continue

            # Skip if already confirmed or declined
            if booking.is_confirmed is not None:
                current_app.logger.info(
                    "Booking for %s %s already has confirmation status, skipping",
                    volunteer.first_name,
                    volunteer.last_name,
                )
                skipped_count += 1
                continue

            # Skip if volunteer was already notified about time change (that message asks for confirmation)
            if booking.time_change_notified:
                current_app.logger.info(
                    "Booking for %s %s already got time-change notification, skipping reminder",
                    volunteer.first_name,
                    volunteer.last_name,
                )
                skipped_count += 1
                continue
            
            # Skip if signed up recently (already got signup confirmation)
            if booking.timestamp:
                ts = booking.timestamp
                if ts.tzinfo is None:
                    ts = ts.replace(tzinfo=timezone.utc)
                if ts > signup_cutoff:
                    current_app.logger.info(
                        "Booking for %s %s is too recent (signed up %s), skipping reminder",
                        volunteer.first_name,
                        volunteer.last_name,
                        booking.timestamp.strftime('%Y-%m-%d %H:%M'),
                    )
                    skipped_count += 1
                    continue

            volunteer_bookings[volunteer.id].append((training, booking))

    # Now send one message per volunteer
    for vol_id, items in volunteer_bookings.items():
        # Sort by training time
        items.sort(key=lambda x: x[0].date)
        volunteer = items[0][1].volunteer
        volunteer_full_name = f"{volunteer.first_name} {volunteer.last_name}"

        if len(items) == 1:
            training, booking = items[0]
            coach_full_name = f"{training.coach.first_name} {training.coach.last_name}"
            success, error = notify_volunteer_reminder(
                volunteer_phone=volunteer.phone_number,
                volunteer_name=volunteer.first_name,
                training_date=training.date.strftime('%Y-%m-%d'),
                training_time=training.date.strftime('%H:%M'),
                training_location=training.location.name,
                coach_name=coach_full_name,
                coach_phone=training.coach.phone_number,
            )
        else:
            trainings_info = []
            for training, booking in items:
                coach_full_name = f"{training.coach.first_name} {training.coach.last_name}"
                trainings_info.append({
                    'time': training.date.strftime('%H:%M'),
                    'location': training.location.name,
                    'coach_name': coach_full_name,
                    'coach_phone': training.coach.phone_number,
                })
            success, error = notify_volunteer_reminder_multi(
                volunteer_phone=volunteer.phone_number,
                volunteer_name=volunteer.first_name,
                trainings_info=trainings_info,
            )

        if success:
            sent_count += 1
            click.echo(f"✓ Reminder sent to {volunteer_full_name} ({len(items)} training(s))")
        else:
            failed_count += 1
            click.echo(f"✗ Failed to send reminder to {volunteer_full_name}: {error}")

    click.echo(f"\nSummary: {sent_count} sent, {failed_count} failed, {skipped_count} skipped")


@click.command('send-phone-requests')
@click.option('--base-url', default='https://treningi.widzimyinaczej.org.pl', 
              help='Base URL for the application')
@with_appcontext
def send_phone_requests_command(base_url):
    """Send email to volunteers without phone numbers asking them to add their phone.
    
    Each volunteer receives this email only once (tracked by phone_request_sent flag).
    """
    # Znajdź wolontariuszy bez telefonu, którzy nie dostali jeszcze maila
    volunteers = Volunteer.query.filter(
        (Volunteer.phone_number.is_(None)) | (Volunteer.phone_number == ''),
        (Volunteer.phone_request_sent.is_(None)) | (Volunteer.phone_request_sent.is_(False)),
    ).all()

    if not volunteers:
        click.echo("No volunteers without phone numbers need to be contacted.")
        return

    click.echo(f"Found {len(volunteers)} volunteers without phone numbers to contact.")

    # Pobierz szablon email
    setting = EmailSettings.query.first()
    if not setting or not setting.phone_request_template:
        click.echo("Error: Phone request email template not configured in admin settings.")
        click.echo("Please go to Admin > Settings and configure the phone request template.")
        return

    template = setting.phone_request_template
    sent_count = 0
    failed_count = 0
    
    # Logo URL
    logo_url = f"{base_url}/static/logo.png"

    for volunteer in volunteers:
        # Generuj unikalny token
        token = secrets.token_urlsafe(32)
        volunteer.phone_update_token = token
        
        # Przygotuj dane do szablonu
        update_link = f"{base_url}/update-phone/{token}"
        data = {
            'first_name': volunteer.first_name,
            'last_name': volunteer.last_name,
            'email': volunteer.email,
            'update_link': update_link,
            'logo': logo_url,
        }
        
        try:
            html_body = render_template_string(template, data)
            success, error = send_email(
                "Dodaj numer telefonu - powiadomienia WhatsApp",
                None,
                [volunteer.email],
                html_body=html_body,
            )
            
            if success:
                # Oznacz że email został wysłany
                volunteer.phone_request_sent = True
                db.session.commit()
                sent_count += 1
                click.echo(f"✓ Email sent to {volunteer.first_name} {volunteer.last_name} ({volunteer.email})")
            else:
                # W razie błędu, nie oznaczaj - spróbuj ponownie następnym razem
                db.session.rollback()
                failed_count += 1
                click.echo(f"✗ Failed to send to {volunteer.email}: {error}")
        except Exception as e:
            db.session.rollback()
            failed_count += 1
            click.echo(f"✗ Error sending to {volunteer.email}: {str(e)}")

    click.echo(f"\nSummary: {sent_count} sent, {failed_count} failed")


@click.command("send-coach-summary")
@click.option("--hours-before", default=1, help="Send summary this many hours before first training")
@click.option("--window-minutes", default=30, help="Time window in minutes (run cron at this interval)")
@with_appcontext
def send_coach_summary_command(hours_before, window_minutes):
    """Send WhatsApp summary to coaches about todays trainings.
    
    Sends to coaches whose first training is approximately --hours-before from now.
    Run this command every --window-minutes via cron.
    
    Example cron (every 30 min): */30 * * * * cd /app && flask send-coach-summary
    """
    from .whatsapp_utils import send_whatsapp_message, format_phone_display
    from .models import Coach

    # Use Warsaw timezone consistently
    warsaw_tz = ZoneInfo("Europe/Warsaw")
    now = datetime.now(warsaw_tz)
    today = now.date()
    
    # Create naive datetimes for DB filtering (assuming DB stores naive datetimes)
    today_start = datetime.combine(today, datetime.min.time())
    today_end = datetime.combine(today, datetime.max.time())

    # Get all trainings for today grouped by coach
    trainings = Training.query.filter(
        Training.date >= today_start,
        Training.date <= today_end,
        Training.is_canceled.is_(False),
        Training.is_deleted.is_(False),
    ).order_by(Training.date).all()

    if not trainings:
        click.echo("No trainings scheduled for today.")
        return

    # Group trainings by coach and find first training time for each
    coach_trainings = {}
    coach_first_training = {}
    for training in trainings:
        coach_id = training.coach_id
        if coach_id not in coach_trainings:
            coach_trainings[coach_id] = []
            # Make training.date timezone-aware (it's stored as naive in DB)
            training_dt = training.date
            if training_dt.tzinfo is None:
                training_dt = training_dt.replace(tzinfo=warsaw_tz)
            coach_first_training[coach_id] = training_dt
        coach_trainings[coach_id].append(training)
        # Track earliest training
        training_dt = training.date
        if training_dt.tzinfo is None:
            training_dt = training_dt.replace(tzinfo=warsaw_tz)
        if training_dt < coach_first_training[coach_id]:
            coach_first_training[coach_id] = training_dt

    sent_count = 0
    failed_count = 0
    skipped_count = 0
    not_yet_count = 0

    for coach_id, coach_trainings_list in coach_trainings.items():
        coach = Coach.query.get(coach_id)
        if not coach or not coach.phone_number:
            current_app.logger.info(
                "Coach %s has no phone number, skipping summary",
                coach.first_name if coach else "Unknown",
            )
            skipped_count += 1
            continue

        # Check if it's time to send (first training is hours_before +/- window_minutes/2)
        first_training_time = coach_first_training[coach_id]
        time_until_training = (first_training_time - now).total_seconds() / 60  # in minutes
        
        target_minutes = hours_before * 60
        window_half = window_minutes / 2
        
        # Send if we're within the window (e.g., 45-75 min before for 1h with 30min window)
        if time_until_training < (target_minutes - window_half):
            # Too late, training is too soon or already started
            current_app.logger.info(
                "Coach %s first training at %s - too late to send (%d min away)",
                coach.first_name,
                first_training_time.strftime("%H:%M"),
                int(time_until_training),
            )
            skipped_count += 1
            continue
        
        if time_until_training > (target_minutes + window_half):
            # Too early, will send later
            current_app.logger.info(
                "Coach %s first training at %s - too early (%d min away, waiting)",
                coach.first_name,
                first_training_time.strftime("%H:%M"),
                int(time_until_training),
            )
            not_yet_count += 1
            continue

        # Skip if coach has NO volunteers at all
        has_any_volunteers = False
        for training in coach_trainings_list:
            if training.bookings:
                has_any_volunteers = True
                break

        if not has_any_volunteers:
            current_app.logger.info(
                "Coach %s has no volunteers today — skipping summary",
                coach.first_name,
            )
            skipped_count += 1
            continue

        message_lines = [
            f"📋 *Dzisiejsze treningi*\n",
            f"Cześć {coach.first_name}! 👋\n",
            "Podsumowanie Twoich dzisiejszych treningów:\n",
        ]

        for training in coach_trainings_list:
            time_str = training.date.strftime("%H:%M")
            location = training.location.name

            # Get confirmed, pending, and declined volunteers
            confirmed_bookings = [b for b in training.bookings if b.is_confirmed is True]
            pending_bookings = [b for b in training.bookings if b.is_confirmed is None]
            declined_bookings = [b for b in training.bookings if b.is_confirmed is False]

            message_lines.append(f"🕐 *{time_str} — {location}*")

            if confirmed_bookings:
                for booking in confirmed_bookings:
                    vol = booking.volunteer
                    if vol.phone_number:
                        phone_str = format_phone_display(vol.phone_number)
                    else:
                        phone_str = "brak tel."
                    message_lines.append(f"  ✅ {vol.first_name} {vol.last_name} ({phone_str})")

            if pending_bookings:
                for booking in pending_bookings:
                    vol = booking.volunteer
                    if vol.phone_number:
                        phone_str = format_phone_display(vol.phone_number)
                    else:
                        phone_str = "brak tel."
                    message_lines.append(f"  ❓ {vol.first_name} {vol.last_name} ({phone_str})")

            if declined_bookings:
                for booking in declined_bookings:
                    vol = booking.volunteer
                    if vol.phone_number:
                        phone_str = format_phone_display(vol.phone_number)
                    else:
                        phone_str = "brak tel."
                    message_lines.append(f"  ❌ {vol.first_name} {vol.last_name} ({phone_str})")

            if not confirmed_bookings and not pending_bookings and not declined_bookings:
                message_lines.append("  ⚪ Brak zapisanych wolontariuszy")

            message_lines.append("")

        message_lines.append("✅ — potwierdzony  ❓ — oczekuje  ❌ — zrezygnował")
        message_lines.append("")
        message_lines.append("🎾 *Fundacja Widzimy Inaczej*\n_System zapisów Blind Tenis_")
        message = "\n".join(message_lines)

        success, error = send_whatsapp_message(coach.phone_number, message)
        coach_name = f"{coach.first_name} {coach.last_name}"

        if success:
            sent_count += 1
            click.echo(f"OK Podsumowanie wyslane do {coach_name} (trening o {first_training_time.strftime('%H:%M')})")
        else:
            failed_count += 1
            click.echo(f"BLAD Nie udalo sie wyslac do {coach_name}: {error}")

    click.echo(f"\nPodsumowanie: {sent_count} wyslanych, {failed_count} bledow, {skipped_count} pominietych, {not_yet_count} za wczesnie")


@click.command("send-monthly-summary")
@click.option("--month", default=None, type=int, help="Month number (1-12). Default: previous month.")
@click.option("--year", default=None, type=int, help="Year. Default: current year (or previous if Jan).")
@click.option("--test-email", default=None, help="Send all summaries to this email instead (for testing).")
@with_appcontext
def send_monthly_summary_command(month, year, test_email):
    """Send monthly training summary email to each coach.

    By default summarises the previous month.  Run on the last day of
    each month (or first day of next month) via cron.

    Example cron (last day of month at 21:00):
      0 21 28-31 * * [ "$(date -d tomorrow +\\%d)" = "01" ] && cd ... && flask send-monthly-summary
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from .models import Coach

    warsaw_tz = ZoneInfo("Europe/Warsaw")
    now = datetime.now(warsaw_tz)

    if month is None or year is None:
        # Default: previous month
        first_of_current = now.replace(day=1)
        last_month = first_of_current - timedelta(days=1)
        month = month or last_month.month
        year = year or last_month.year

    # Month date range (naive, matching DB storage)
    month_start = datetime(year, month, 1)
    if month == 12:
        month_end = datetime(year + 1, 1, 1)
    else:
        month_end = datetime(year, month + 1, 1)

    MONTH_NAMES_PL = [
        '', 'styczeń', 'luty', 'marzec', 'kwiecień', 'maj', 'czerwiec',
        'lipiec', 'sierpień', 'wrzesień', 'październik', 'listopad', 'grudzień',
    ]
    month_name = MONTH_NAMES_PL[month]
    period_label = f"{month_name} {year}"

    coaches = Coach.query.order_by(Coach.last_name).all()
    sent_count = 0
    skipped_count = 0
    failed_count = 0

    for coach in coaches:
        recipient = test_email or coach.email
        if not recipient:
            click.echo(f"POMINIĘTO {coach.first_name} {coach.last_name} — brak emaila")
            skipped_count += 1
            continue

        # Get this coach's trainings for the month
        trainings = (
            Training.query
            .filter(
                Training.coach_id == coach.id,
                Training.date >= month_start,
                Training.date < month_end,
                Training.is_canceled.is_(False),
                Training.is_deleted.is_(False),
            )
            .order_by(Training.date)
            .all()
        )

        if not trainings:
            click.echo(f"POMINIĘTO {coach.first_name} {coach.last_name} — brak treningów w {period_label}")
            skipped_count += 1
            continue

        # ── Build Excel workbook ────────────────────────────────
        wb = Workbook()
        ws = wb.active
        ws.title = f"Treningi {period_label}"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        # Title row
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = f"Podsumowanie treningów — {period_label}"
        title_cell.font = Font(bold=True, size=14, color="2E7D32")
        title_cell.alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:D2')
        ws['A2'].value = f"Trener: {coach.first_name} {coach.last_name}"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = Alignment(horizontal='center')

        # Headers
        headers = ['Data', 'Godzina', 'Miejsce', 'Wolontariusze']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        row = 5
        total_trainings = 0
        total_volunteers = 0

        for training in trainings:
            total_trainings += 1
            volunteers = [b.volunteer for b in training.bookings]
            vol_names = ', '.join(
                f"{v.first_name} {v.last_name}" for v in volunteers
            ) if volunteers else '—'
            total_volunteers += len(volunteers)

            date_str = training.date.strftime('%d.%m.%Y')
            time_str = training.date.strftime('%H:%M')
            loc_name = training.location.name

            ws.cell(row=row, column=1, value=date_str).border = thin_border
            ws.cell(row=row, column=2, value=time_str).border = thin_border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=loc_name).border = thin_border
            ws.cell(row=row, column=4, value=vol_names).border = thin_border
            row += 1

        # Summary row
        row += 1
        ws.cell(row=row, column=1, value='Łącznie treningów:').font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_trainings).font = Font(bold=True)
        ws.cell(row=row, column=1 + 2, value='Łącznie wolontariuszy:').font = Font(bold=True)
        ws.cell(row=row, column=4, value=total_volunteers).font = Font(bold=True)

        # Auto-width
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 45

        buf = io.BytesIO()
        wb.save(buf)
        excel_bytes = buf.getvalue()
        filename = f"treningi_{month_name}_{year}.xlsx"

        # ── Build HTML email body ───────────────────────────────
        html_rows = []
        for training in trainings:
            volunteers = [b.volunteer for b in training.bookings]
            vol_str = ', '.join(
                f"{v.first_name} {v.last_name}" for v in volunteers
            ) if volunteers else '<em>brak</em>'
            html_rows.append(
                f"<tr>"
                f"<td style='padding:6px 10px;border:1px solid #ddd;'>{training.date.strftime('%d.%m.%Y')}</td>"
                f"<td style='padding:6px 10px;border:1px solid #ddd;text-align:center;'>{training.date.strftime('%H:%M')}</td>"
                f"<td style='padding:6px 10px;border:1px solid #ddd;'>{training.location.name}</td>"
                f"<td style='padding:6px 10px;border:1px solid #ddd;'>{vol_str}</td>"
                f"</tr>"
            )

        html_body = f"""\
<div style="font-family:Arial,sans-serif;max-width:700px;margin:0 auto;">
  <h2 style="color:#2E7D32;">🎾 Podsumowanie treningów — {period_label}</h2>
  <p>Cześć {coach.first_name}!</p>
  <p>Poniżej znajdziesz podsumowanie Twoich treningów za <strong>{period_label}</strong>:</p>
  <table style="border-collapse:collapse;width:100%;margin:16px 0;">
    <thead>
      <tr style="background:#2E7D32;color:#fff;">
        <th style="padding:8px 10px;border:1px solid #2E7D32;">Data</th>
        <th style="padding:8px 10px;border:1px solid #2E7D32;">Godzina</th>
        <th style="padding:8px 10px;border:1px solid #2E7D32;">Miejsce</th>
        <th style="padding:8px 10px;border:1px solid #2E7D32;">Wolontariusze</th>
      </tr>
    </thead>
    <tbody>
      {''.join(html_rows)}
    </tbody>
  </table>
  <p><strong>Łącznie:</strong> {total_trainings} treningów, {total_volunteers} wolontariuszy</p>
  <p style="margin-top:24px;color:#666;font-size:13px;">
    Ten email został wygenerowany automatycznie.<br>
    🎾 Fundacja Widzimy Inaczej — Blind Tenis
  </p>
</div>"""

        # ── Send email ──────────────────────────────────────────
        success, error = send_email(
            subject=f"Podsumowanie treningów — {period_label}",
            body=None,
            recipients=[recipient],
            html_body=html_body,
            attachments=[(filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", excel_bytes)],
        )

        coach_name = f"{coach.first_name} {coach.last_name}"
        if success:
            sent_count += 1
            dest = f" → {recipient}" if test_email else ""
            click.echo(f"OK {coach_name}: {total_trainings} treningów, {total_volunteers} wolontariuszy{dest}")
        else:
            failed_count += 1
            click.echo(f"BŁĄD {coach_name}: {error}")

    click.echo(f"\nWysłano: {sent_count}, pominięto: {skipped_count}, błędów: {failed_count}")


def init_app(app):
    """Register CLI commands with the app."""
    app.cli.add_command(send_reminders_command)
    app.cli.add_command(send_phone_requests_command)
    app.cli.add_command(send_coach_summary_command)
    app.cli.add_command(send_monthly_summary_command)
